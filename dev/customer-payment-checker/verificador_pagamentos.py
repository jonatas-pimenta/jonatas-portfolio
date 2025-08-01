"""
Sistema de verificação automática de pagamentos de clientes via consulta CPF.

Módulo principal do Customer Payment Checker que processa planilhas Excel
e consulta status de pagamentos em site externo com múltiplas opções de processamento.

Uso: ./executar.sh (recomendado) ou python verificador_pagamentos.py
"""

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
import os

def configurar_driver():
    """Configura o driver do Chrome com opções otimizadas"""
    chrome_options = Options()
    # chrome_options.add_argument("--headless")  # Descomente para executar sem interface gráfica
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    
    # Tentar usar chromedriver do sistema primeiro
    try:
        service = Service("/usr/bin/chromedriver")
        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver
    except:
        # Fallback para webdriver-manager
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver

def criar_planilha_exemplo():
    """Cria planilha de exemplo com dados fictícios para teste"""
    print("🏗️  Criando planilha de exemplo...")
    
    # Criar nova planilha
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'
    
    # Cabeçalho
    sheet.append(["Nome", "Valor", "CPF", "Vencimento"])
    
    # Dados de exemplo (CPFs fictícios para teste)
    dados_exemplo = [
        ["João Silva", 1500.00, "12345678901", "2024-01-15"],
        ["Maria Santos", 2300.50, "98765432100", "2024-01-20"],
        ["Carlos Oliveira", 850.75, "11122233344", "2024-01-25"],
        ["Ana Costa", 1200.00, "55566677788", "2024-02-01"],
        ["Pedro Souza", 950.25, "99988877766", "2024-02-05"]
    ]
    
    # Adicionar dados
    for linha in dados_exemplo:
        sheet.append(linha)
    
    # Salvar
    caminho = "dados_clientes_exemplo.xlsx"
    workbook.save(caminho)
    
    print(f"✅ Planilha criada: {caminho}")
    print(f"📊 {len(dados_exemplo)} clientes de exemplo")
    return caminho

def verificar_planilha_clientes(caminho_planilha):
    """Verifica se a planilha de clientes existe"""
    if not os.path.exists(caminho_planilha):
        print(f"Erro: Planilha de clientes não encontrada em {caminho_planilha}")
        return False
    return True

def criar_ou_carregar_planilha_fechamento(caminho_fechamento):
    """Cria ou carrega a planilha de fechamento"""
    try:
        planilha_fechamento = openpyxl.load_workbook(caminho_fechamento)
        pagina_fechamento = planilha_fechamento['Sheet1']
        print("Planilha de fechamento carregada com sucesso!")
    except FileNotFoundError:
        planilha_fechamento = openpyxl.Workbook()
        pagina_fechamento = planilha_fechamento.active
        pagina_fechamento.title = 'Sheet1'
        pagina_fechamento.append(["Nome", "Valor", "CPF", "Vencimento", "Status", "Data do Pagamento", "Método de Pagamento"])
        planilha_fechamento.save(caminho_fechamento)
        print("Planilha de fechamento criada com sucesso!")
    
    return planilha_fechamento, pagina_fechamento

def pesquisar_status_cliente(driver, cpf):
    """Pesquisa o status do cliente no site"""
    try:
        # Aguardar a página carregar
        sleep(2)
        
        # Preencher o CPF no campo de pesquisa
        campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
        campo_pesquisa.clear()
        campo_pesquisa.send_keys(cpf)
        sleep(1)

        # Clicar no botão de pesquisa
        botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
        botao_pesquisar.click()
        sleep(4)

        # Capturar o status do cliente
        status_element = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
        status = status_element.text.strip()

        if status.lower() == 'em dia':
            # Capturar data do pagamento e método de pagamento
            data_pagamento_element = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
            metodo_pagamento_element = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")

            data_pagamento = data_pagamento_element.text.split(":")[1].strip()
            metodo_pagamento = metodo_pagamento_element.text.split(":")[1].strip()

            return status, data_pagamento, metodo_pagamento
        else:
            return status, None, None
            
    except Exception as e:
        print(f"Erro ao pesquisar CPF {cpf}: {e}")
        return "Erro", None, None

def mostrar_preview_clientes(pagina_clientes, limite=10):
    """Mostra uma prévia dos clientes na planilha"""
    print("\n📋 Prévia dos clientes na planilha:")
    print("-" * 60)
    count = 0
    for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:
            continue
        count += 1
        nome, valor, cpf, vencimento = linha
        print(f"{count:2d}. {nome} (CPF: {cpf}) - R$ {valor}")
        if count >= limite:
            break
    
    total = sum(1 for row in pagina_clientes.iter_rows(min_row=2, values_only=True) if row[0] is not None)
    if total > limite:
        print(f"    ... e mais {total - limite} clientes")
    print(f"\n📊 Total de clientes na planilha: {total}")
    return total

def obter_opcoes_processamento(total_clientes):
    """Permite ao usuário escolher como processar os clientes"""
    print("\n🔧 Opções de processamento:")
    print("1. Processar TODOS os clientes")
    print("2. Processar apenas os primeiros N clientes")
    print("3. Processar em lotes (com pausas)")
    print("4. Processar um intervalo específico (ex: do 5º ao 15º cliente)")
    print("5. Continuar de onde parou (se já processou alguns)")
    
    while True:
        try:
            opcao = input("\nEscolha uma opção (1-5): ").strip()
            
            if opcao == "1":
                return "todos", total_clientes, 0
            
            elif opcao == "2":
                quantidade = int(input(f"Quantos clientes processar? (máximo {total_clientes}): "))
                if 1 <= quantidade <= total_clientes:
                    return "primeiros", quantidade, 0
                else:
                    print(f"❌ Digite um número entre 1 e {total_clientes}")
            
            elif opcao == "3":
                tamanho_lote = int(input("Tamanho de cada lote: "))
                if 1 <= tamanho_lote <= total_clientes:
                    return "lotes", tamanho_lote, 0
                else:
                    print(f"❌ Digite um número entre 1 e {total_clientes}")
            
            elif opcao == "4":
                inicio = int(input(f"Começar do cliente número: "))
                fim = int(input(f"Terminar no cliente número: "))
                if 1 <= inicio <= fim <= total_clientes:
                    return "intervalo", fim - inicio + 1, inicio - 1
                else:
                    print(f"❌ Intervalo inválido. Use números entre 1 e {total_clientes}")
            
            elif opcao == "5":
                inicio = int(input(f"Continuar a partir do cliente número: "))
                if 1 <= inicio <= total_clientes:
                    return "continuar", total_clientes - inicio + 1, inicio - 1
                else:
                    print(f"❌ Digite um número entre 1 e {total_clientes}")
            
            else:
                print("❌ Opção inválida. Digite 1, 2, 3, 4 ou 5")
                
        except ValueError:
            print("❌ Digite apenas números")
        except KeyboardInterrupt:
            print("\n❌ Operação cancelada pelo usuário")
            return None, 0, 0

def processar_clientes_com_controle(driver, pagina_clientes, pagina_fechamento, planilha_fechamento, caminho_fechamento, modo, quantidade, inicio_idx):
    """Processa os clientes conforme as opções escolhidas"""
    clientes_processados = 0
    cliente_atual = 0
    
    # Pular clientes até o índice de início
    clientes_lista = []
    for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
        if linha[0] is not None:
            clientes_lista.append(linha)
    
    clientes_para_processar = clientes_lista[inicio_idx:]
    
    if modo == "lotes":
        tamanho_lote = quantidade
        total_lotes = (len(clientes_para_processar) + tamanho_lote - 1) // tamanho_lote
        print(f"\n📦 Processamento em lotes de {tamanho_lote} clientes")
        print(f"Total de lotes: {total_lotes}")
        
        for lote_idx in range(total_lotes):
            inicio_lote = lote_idx * tamanho_lote
            fim_lote = min(inicio_lote + tamanho_lote, len(clientes_para_processar))
            
            print(f"\n🔄 Processando lote {lote_idx + 1}/{total_lotes}")
            
            for i in range(inicio_lote, fim_lote):
                linha = clientes_para_processar[i]
                cliente_atual = inicio_idx + i + 1
                processar_cliente_individual(driver, linha, pagina_fechamento, planilha_fechamento, caminho_fechamento, cliente_atual, len(clientes_lista))
                clientes_processados += 1
            
            if lote_idx < total_lotes - 1:  # Não pausar no último lote
                print(f"✅ Lote {lote_idx + 1} concluído!")
                resposta = input("Pressione Enter para continuar com o próximo lote ou 'q' para parar: ")
                if resposta.lower() == 'q':
                    break
    
    else:
        limite = min(quantidade, len(clientes_para_processar))
        for i in range(limite):
            linha = clientes_para_processar[i]
            cliente_atual = inicio_idx + i + 1
            processar_cliente_individual(driver, linha, pagina_fechamento, planilha_fechamento, caminho_fechamento, cliente_atual, len(clientes_lista))
            clientes_processados += 1
    
    return clientes_processados

def processar_cliente_individual(driver, linha, pagina_fechamento, planilha_fechamento, caminho_fechamento, cliente_atual, total_clientes):
    """Processa um cliente individual"""
    nome, valor, cpf, vencimento = linha
    print(f"Processando cliente {cliente_atual}/{total_clientes}: {nome} (CPF: {cpf})")

    # Pesquisar status do cliente
    status, data_pagamento, metodo_pagamento = pesquisar_status_cliente(driver, cpf)

    # Processar resultado baseado no status
    if status.lower() == 'em dia':
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'Em dia', data_pagamento, metodo_pagamento])
        print(f"  ✓ Cliente em dia - Pagamento: {data_pagamento} via {metodo_pagamento}")
    elif status.lower() == 'atrasado':
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'Pendente', None, None])
        print(f"  ⚠ Cliente com pagamento pendente")
    else:
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'Erro na consulta', None, None])
        print(f"  ❌ Erro ao consultar cliente")

    # Salvar a planilha após cada cliente processado
    planilha_fechamento.save(caminho_fechamento)

def main():
    """Função principal do programa"""
    # Caminhos dos arquivos
    workspace_path = "/home/lion/Documentos/Projetos/jonatas-portfolio/dev/customer-payment-checker"
    caminho_planilha_clientes = os.path.join(workspace_path, "dados_clientes.xlsx")
    caminho_planilha_fechamento = os.path.join(workspace_path, "planilha_fechamento.xlsx")
    
    # Verificar se a planilha de clientes existe
    if not verificar_planilha_clientes(caminho_planilha_clientes):
        return
    
    # Configurar o driver
    driver = None
    try:
        # Carregar a planilha de clientes primeiro para mostrar prévia
        print("Carregando planilha de clientes...")
        planilha_clientes = openpyxl.load_workbook(caminho_planilha_clientes)
        pagina_clientes = planilha_clientes['Sheet1']

        # Mostrar prévia dos clientes
        total_clientes = mostrar_preview_clientes(pagina_clientes)
        
        # Obter opções de processamento do usuário
        modo, quantidade, inicio_idx = obter_opcoes_processamento(total_clientes)
        
        if modo is None:  # Usuário cancelou
            return
        
        print(f"\n🚀 Iniciando processamento...")
        if modo == "todos":
            print(f"📊 Processando TODOS os {total_clientes} clientes")
        elif modo == "primeiros":
            print(f"📊 Processando os primeiros {quantidade} clientes")
        elif modo == "lotes":
            print(f"📊 Processando em lotes de {quantidade} clientes")
        elif modo == "intervalo":
            print(f"📊 Processando clientes do {inicio_idx + 1}º ao {inicio_idx + quantidade}º")
        elif modo == "continuar":
            print(f"📊 Continuando do {inicio_idx + 1}º cliente até o final")

        print("Iniciando o navegador...")
        driver = configurar_driver()
        driver.get('https://consultcpf-devaprender.netlify.app/')
        print("Site carregado com sucesso!")

        # Criar ou carregar a planilha de fechamento
        planilha_fechamento, pagina_fechamento = criar_ou_carregar_planilha_fechamento(caminho_planilha_fechamento)

        # Processar clientes conforme as opções escolhidas
        clientes_processados = processar_clientes_com_controle(
            driver, pagina_clientes, pagina_fechamento, 
            planilha_fechamento, caminho_planilha_fechamento,
            modo, quantidade, inicio_idx
        )

        print(f"\n✅ Processo concluído! {clientes_processados} clientes processados.")
        print(f"Dados salvos em: {caminho_planilha_fechamento}")

    except Exception as e:
        print(f"❌ Erro durante o processo: {e}")
    finally:
        # Garantir que o navegador seja fechado
        if driver:
            print("Fechando navegador...")
            driver.quit()
            print("Navegador fechado com sucesso!")

if __name__ == "__main__":
    main()
